import logging
import os
import time
from datetime import datetime, timedelta, timezone

from dotenv import load_dotenv
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    make_response,
    jsonify,
    send_file,
)
from flask_wtf.csrf import CSRFProtect
from sqlalchemy import inspect
from sqlalchemy.exc import OperationalError, TimeoutError
from werkzeug.middleware.proxy_fix import ProxyFix

from utils.timezone import KKTC_TZ, get_kktc_now

# Logging ayarla - Hem console hem de dosyaya yaz
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("app.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

# .env dosyasını yükle
load_dotenv()

# Flask uygulaması oluştur
app = Flask(__name__)

# Konfigürasyonu yükle
app.config.from_object("config.Config")

# Proxy arkasında çalışırken (Traefik/Nginx) doğru scheme ve IP algılama
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# CSRF Koruması Aktif
csrf = CSRFProtect(app)


# CSRF token'ı tüm template'lere ekle
@app.context_processor
def inject_csrf_token():
    """CSRF token'ı template'lere enjekte et"""
    from flask_wtf.csrf import generate_csrf

    # Hem fonksiyon hem de değişken olarak sağla (geriye dönük uyumluluk için)
    return dict(csrf_token=generate_csrf)


# Rate Limiting Devre Dışı (İhtiyaç halinde açılabilir)
# limiter = Limiter(
#     app=app,
#     key_func=get_remote_address,
#     default_limits=["200 per day", "50 per hour"],
#     storage_uri="memory://",  # Production'da Redis kullanılmalı
#     strategy="fixed-window"
# )

# Veritabanı başlat
from models import db  # noqa: E402
from flask_migrate import Migrate  # noqa: E402

db.init_app(app)
migrate = Migrate(app, db)

# Veritabanı metadata'sını yenile ve bağlantıyı test et
with app.app_context():
    try:
        db.engine.dispose()
        db.reflect()
        # Bağlantıyı test et
        with db.engine.connect() as conn:
            result = conn.execute(db.text("SELECT 1"))
            result.close()
        logger.info(
            "✅ Database engine yenilendi, metadata reflect edildi ve bağlantı test edildi"
        )
    except Exception as e:
        logger.warning(f"⚠️ Database başlatma hatası (görmezden geliniyor): {e}")

# Cache devre dışı (Redis sadece Celery broker olarak kullanılıyor)
cache = None
logger.info("ℹ️ Cache devre dışı - Redis sadece Celery broker olarak kullanılıyor")

# ============================================
# RATE LIMITER INITIALIZATION
# ============================================
limiter = None
if app.config.get("RATE_LIMIT_ENABLED", True):
    try:
        from utils.rate_limiter import init_rate_limiter

        limiter = init_rate_limiter(app)
        logger.info("✅ Rate Limiter aktifleştirildi")
    except Exception as e:
        logger.warning(f"⚠️ Rate Limiter başlatılamadı: {str(e)}")
else:
    logger.info("ℹ️ Rate Limiter devre dışı (config)")

# ============================================
# CACHE MANAGER INITIALIZATION (Master Data Only)
# ============================================
cache_manager = None
if app.config.get("CACHE_ENABLED", True):
    try:
        from utils.cache_manager import init_cache

        cache_manager = init_cache(app)
        logger.info("✅ Cache Manager aktifleştirildi (sadece master data)")
    except Exception as e:
        logger.warning(f"⚠️ Cache Manager başlatılamadı: {str(e)}")

# Query Logging - SQLAlchemy Event Listener
try:
    from utils.monitoring.query_analyzer import setup_query_logging

    setup_query_logging()
except Exception as e:
    logger.warning(f"Query logging setup hatası: {e}")

# İkinci dispose kaldırıldı - yukarıda zaten yapılıyor (10.02.2026)
# Engine test yukarıdaki blokta yapılıyor


# Database Connection Retry Mekanizması - Railway Timeout Fix v3 (ULTRA AGRESIF)
def init_db_with_retry(max_retries=3, retry_delay=10):
    """
    Database bağlantısını retry mekanizması ile başlat
    Railway'de cold start veya network timeout sorunlarını çözer
    v3: Daha uzun timeout, daha az deneme
    """
    for attempt in range(max_retries):
        try:
            with app.app_context():
                # Database bağlantısını test et
                connection = db.engine.connect()
                connection.close()
                logger.info(
                    f"✅ Database bağlantısı başarılı (Deneme {attempt + 1}/{max_retries})"
                )
                return True
        except (OperationalError, TimeoutError) as e:
            error_msg = str(e)
            logger.warning(
                f"⚠️ Database bağlantı hatası (Deneme {attempt + 1}/{max_retries}): {error_msg[:200]}"
            )

            if attempt < max_retries - 1:
                # Sabit 30 saniye bekleme (exponential backoff yerine)
                wait_time = 30
                logger.info(f"🔄 {wait_time} saniye sonra tekrar denenecek...")
                time.sleep(wait_time)
            else:
                logger.error(
                    f"❌ Database bağlantısı {max_retries} denemeden sonra başarısız!"
                )
                logger.error(f"❌ Son hata: {error_msg}")
                # Production'da uygulama çalışmaya devam etsin
                return False
        except Exception as e:
            logger.error(f"❌ Beklenmeyen hata: {str(e)}")
            # Beklenmeyen hatalarda da devam et
            return False
    return False


# Uygulama başlatıldığında database bağlantısını test et
try:
    init_db_with_retry()
except Exception as e:
    logger.error(f"❌ FATAL: Database başlatılamadı: {str(e)}")
    # Production'da uygulama çalışmaya devam etsin, ilk request'te tekrar denenecek

# Yardımcı modülleri import et
from utils.decorators import login_required, role_required  # noqa: E402
from utils.helpers import (  # noqa: E402
    get_current_user,
)

# Modelleri import et
from models import (  # noqa: E402
    Kullanici,
    MinibarIslem,
    MinibarIslemDetay,
    SistemLog,
    MinibarDolumTalebi,
)


# Context processor - tüm template'lere kullanıcı bilgisini gönder
@app.context_processor
def inject_user():
    return dict(current_user=get_current_user())


# Context processor - Python built-in fonksiyonları
@app.context_processor
def inject_builtins():
    return dict(min=min, max=max)


# Context processor - Cache version
@app.context_processor
def inject_cache_version():
    """Cache busting için version numarası"""
    from config import Config

    return dict(cache_version=Config.CACHE_VERSION)


# Context processor - Datetime ve tarih fonksiyonları
@app.context_processor
def inject_datetime():
    """Şablonlara datetime ve tarih yardımcı fonksiyonlarını ekle"""
    gun_adlari = [
        "Pazartesi",
        "Salı",
        "Çarşamba",
        "Perşembe",
        "Cuma",
        "Cumartesi",
        "Pazar",
    ]
    return dict(now=datetime.now, gun_adlari=gun_adlari)


# Context processor - Otel bilgisi ve logo
@app.context_processor
def inject_otel_info():
    """Kullanıcının otel bilgisini ve logosunu template'lere gönder.

    flask.g request-scoped cache kullanır - aynı request içinde birden fazla
    çağrılsa bile DB'ye sadece bir kez sorgu gönderilir.
    """
    from flask import g

    if not hasattr(g, '_kullanici_otel'):
        from models import Otel
        from utils.authorization import get_kat_sorumlusu_otel, get_depo_sorumlusu_oteller

        kullanici = get_current_user()
        otel_bilgi = None

        if kullanici:
            try:
                if kullanici.rol == "kat_sorumlusu":
                    otel = get_kat_sorumlusu_otel(kullanici.id)
                    if otel:
                        otel_bilgi = {"ad": otel.ad, "logo": otel.logo}

                elif kullanici.rol == "depo_sorumlusu":
                    oteller = get_depo_sorumlusu_oteller(kullanici.id)
                    if oteller:
                        otel_bilgi = {"ad": oteller[0].ad, "logo": oteller[0].logo}

                elif kullanici.rol in ["admin", "sistem_yoneticisi"]:
                    otel = Otel.query.filter_by(aktif=True).first()
                    if otel:
                        otel_bilgi = {"ad": otel.ad, "logo": otel.logo}
            except Exception as e:
                logger.error(f"Otel bilgisi alınırken hata: {str(e)}")

        g._kullanici_otel = otel_bilgi

    return dict(kullanici_otel=g._kullanici_otel)


# ============================================
# CACHE CONTROL - Template ve HTML cache'ini devre dışı bırak
# ============================================
@app.after_request
def add_no_cache_headers(response):
    """HTML response'larına no-cache header'ları ekle"""
    # Sadece HTML sayfaları için cache'i devre dışı bırak
    if response.content_type and "text/html" in response.content_type:
        response.headers["Cache-Control"] = (
            "no-cache, no-store, must-revalidate, max-age=0"
        )
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


# ============================================
# PWA SUPPORT - Service Worker
# ============================================
@app.route("/sw.js")
def service_worker():
    """Service Worker dosyasını root'tan serve et"""
    return send_file("static/sw.js", mimetype="application/javascript")


# ============================================
# METRICS MIDDLEWARE
# ============================================
from middleware.metrics_middleware import init_metrics_middleware  # noqa: E402

init_metrics_middleware(app)

# ============================================
# ROUTE REGISTRATION - Merkezi Route Yönetimi
# ============================================
from routes import register_all_routes  # noqa: E402

register_all_routes(app)

# ============================================
# Route'lar artık modüler dosyalarda
# Zimmet: routes/zimmet_routes.py
# Minibar Admin: routes/admin_minibar_routes.py
# Depo Raporlar: routes/rapor_routes.py
# ============================================

# Eski API endpoint'leri (artık routes/api_routes.py'de):
# - /api/odalar
# - /api/odalar-by-kat/<int:kat_id>
# - /api/urun-gruplari
# - /api/urunler
# - /api/urunler-by-grup/<int:grup_id>
# - /api/stok-giris
# - /api/minibar-islem-kaydet
# - /api/minibar-ilk-dolum
# - /api/minibar-ilk-dolum-kontrol/<int:oda_id>
# - /api/urun-stok/<int:urun_id>
# - /api/zimmetim
# - /api/minibar-icerigi/<int:oda_id>
# - /api/minibar-doldur
# - /api/toplu-oda-mevcut-durum
# - /api/toplu-oda-doldur
# - /api/kat-rapor-veri


# Kullanım Kılavuzu Sayfası
@app.route("/kullanim-kilavuzu/personel-zimmet")
@login_required
def kullanim_kilavuzu_personel_zimmet():
    """Personel zimmet kullanım kılavuzu sayfası"""
    return render_template("kullanim_kilavuzu/personel_zimmet_kilavuzu.html")


# Dolum Talepleri Rotaları
# ============================================================================
# TOPLU İŞLEM ÖZELLİKLERİ
# ============================================================================


# Excel Export
@app.route("/excel-export/<rapor_tipi>")
@login_required
def excel_export(rapor_tipi):
    try:
        from utils.rapor_export_service import (
            generate_excel,
            get_export_filename,
            VALID_RAPOR_TIPLERI,
        )

        if rapor_tipi not in VALID_RAPOR_TIPLERI:
            flash("Gecersiz rapor tipi", "danger")
            return redirect(url_for("depo_raporlar"))

        filters = {
            "baslangic_tarihi": request.args.get("baslangic_tarihi"),
            "bitis_tarihi": request.args.get("bitis_tarihi"),
            "urun_grup_id": request.args.get("urun_grup_id"),
            "urun_id": request.args.get("urun_id"),
            "personel_id": request.args.get("personel_id"),
            "hareket_tipi": request.args.get("hareket_tipi"),
        }

        output = generate_excel(rapor_tipi, filters)
        filename = get_export_filename(rapor_tipi, "xlsx")

        response = make_response(output.getvalue())
        response.headers["Content-Type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    except Exception as e:
        logger.error("Excel export hatasi: %s", str(e), exc_info=True)
        flash("Sunucu hatasi olustu", "danger")
        return redirect(url_for("depo_raporlar"))


# PDF Export
@app.route("/pdf-export/<rapor_tipi>")
@login_required
def pdf_export(rapor_tipi):
    try:
        from utils.rapor_export_service import (
            generate_pdf,
            get_export_filename,
            VALID_RAPOR_TIPLERI,
        )

        if rapor_tipi not in VALID_RAPOR_TIPLERI:
            flash("Gecersiz rapor tipi", "danger")
            return redirect(url_for("depo_raporlar"))

        filters = {
            "baslangic_tarihi": request.args.get("baslangic_tarihi"),
            "bitis_tarihi": request.args.get("bitis_tarihi"),
            "urun_grup_id": request.args.get("urun_grup_id"),
            "urun_id": request.args.get("urun_id"),
            "personel_id": request.args.get("personel_id"),
            "hareket_tipi": request.args.get("hareket_tipi"),
        }

        output = generate_pdf(rapor_tipi, filters)
        filename = get_export_filename(rapor_tipi, "pdf")

        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "application/pdf"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    except Exception as e:
        logger.error("PDF export hatasi: %s", str(e), exc_info=True)
        flash("Sunucu hatasi olustu", "danger")
        return redirect(url_for("depo_raporlar"))


# ============================================================================
# API: DASHBOARD WIDGET'LARI
# ============================================================================


@app.route("/api/son-aktiviteler")
@login_required
@role_required(["sistem_yoneticisi", "admin"])
def api_son_aktiviteler():
    """Son kullanıcı aktivitelerini döndür"""
    try:
        limit = request.args.get("limit", 10, type=int)

        # Son aktiviteleri çek (sadece önemli işlemler, superadmin hariç)
        aktiviteler = (
            SistemLog.query.join(Kullanici, SistemLog.kullanici_id == Kullanici.id)
            .filter(
                SistemLog.islem_tipi.in_(["ekleme", "guncelleme", "silme"]),
                Kullanici.rol != "superadmin",
            )
            .order_by(SistemLog.islem_tarihi.desc())
            .limit(limit)
            .all()
        )

        data = []
        for log in aktiviteler:
            try:
                # Kullanıcı bilgisi
                kullanici_adi = "Sistem"
                if log.kullanici:
                    kullanici_adi = f"{log.kullanici.ad} {log.kullanici.soyad}"

                # İşlem detayını parse et
                import json

                detay = {}
                if log.islem_detay:
                    try:
                        detay = (
                            json.loads(log.islem_detay)
                            if isinstance(log.islem_detay, str)
                            else log.islem_detay
                        )
                    except Exception:
                        detay = {"aciklama": str(log.islem_detay)}

                # Zaman farkı hesapla
                # islem_tarihi'ni datetime'a çevir
                if isinstance(log.islem_tarihi, datetime):
                    # Datetime objesi
                    if log.islem_tarihi.tzinfo is None:
                        # Naive datetime ise, UTC olarak kabul et
                        islem_tarihi = log.islem_tarihi.replace(tzinfo=timezone.utc)
                    else:
                        islem_tarihi = log.islem_tarihi
                else:
                    # Date objesi ise datetime'a çevir
                    islem_tarihi = datetime.combine(
                        log.islem_tarihi, datetime.min.time()
                    ).replace(tzinfo=timezone.utc)

                zaman_farki = get_kktc_now() - islem_tarihi

                if zaman_farki < timedelta(minutes=1):
                    zaman_str = "Az önce"
                elif zaman_farki < timedelta(hours=1):
                    dakika = int(zaman_farki.total_seconds() / 60)
                    zaman_str = f"{dakika} dakika önce"
                elif zaman_farki < timedelta(days=1):
                    saat = int(zaman_farki.total_seconds() / 3600)
                    zaman_str = f"{saat} saat önce"
                else:
                    gun = zaman_farki.days
                    zaman_str = f"{gun} gün önce"

                data.append(
                    {
                        "id": log.id,
                        "kullanici": kullanici_adi,
                        "islem_tipi": log.islem_tipi,
                        "modul": log.modul,
                        "detay": detay,
                        "zaman": zaman_str,
                        "tam_tarih": islem_tarihi.strftime("%d.%m.%Y %H:%M"),
                    }
                )
            except Exception as log_error:
                # Tek bir log hatası tüm endpoint'i bozmasın
                print(f"Log parse hatası (ID: {log.id}): {log_error}")
                continue

        return jsonify({"success": True, "aktiviteler": data})

    except Exception as e:
        print(f"Son aktiviteler hatası: {e}")
        import traceback

        traceback.print_exc()
        return jsonify({"success": False, "error": "Sunucu hatasi olustu"}), 500


@app.route("/api/bekleyen-dolum-sayisi")
@login_required
@role_required(["sistem_yoneticisi", "admin", "depo_sorumlusu", "kat_sorumlusu"])
def api_bekleyen_dolum_sayisi():
    """Bekleyen dolum talepleri sayısını döndür"""
    try:
        # Bekleyen dolum taleplerini say
        count = MinibarDolumTalebi.query.filter_by(durum="beklemede").count()

        return jsonify({"success": True, "count": count})
    except Exception as e:
        logger.error(f"Bekleyen dolum sayısı hatası: {e}")
        return jsonify(
            {"success": False, "count": 0, "error": "Sunucu hatasi olustu"}
        ), 500


@app.route("/api/tuketim-trendleri")
@login_required
@role_required(["sistem_yoneticisi", "admin", "depo_sorumlusu"])
def api_tuketim_trendleri():
    """Günlük/haftalık tüketim trendlerini döndür"""
    try:
        from sqlalchemy import func

        gun_sayisi = request.args.get("gun", 7, type=int)  # Varsayılan 7 gün

        # Son N günün tüketim verilerini al
        baslangic = get_kktc_now() - timedelta(days=gun_sayisi)

        # Günlük tüketim toplamı (MinibarIslemDetay'dan)
        gunluk_tuketim = (
            db.session.query(
                func.date(MinibarIslem.islem_tarihi).label("tarih"),
                func.sum(MinibarIslemDetay.tuketim).label("toplam_tuketim"),
                func.count(MinibarIslemDetay.id).label("islem_sayisi"),
            )
            .join(MinibarIslemDetay, MinibarIslemDetay.islem_id == MinibarIslem.id)
            .filter(MinibarIslem.islem_tarihi >= baslangic)
            .group_by(func.date(MinibarIslem.islem_tarihi))
            .order_by(func.date(MinibarIslem.islem_tarihi))
            .all()
        )

        # Tüm günleri doldur (veri olmayan günler için 0)
        tum_gunler = {}
        for i in range(gun_sayisi):
            tarih = (get_kktc_now() - timedelta(days=gun_sayisi - i - 1)).date()
            tum_gunler[str(tarih)] = {"tuketim": 0, "islem_sayisi": 0}

        # Veri olanları güncelle
        for row in gunluk_tuketim:
            tarih_str = str(row.tarih)
            tum_gunler[tarih_str] = {
                "tuketim": int(row.toplam_tuketim or 0),
                "islem_sayisi": int(row.islem_sayisi or 0),
            }

        # Chart.js formatına çevir
        labels = []
        tuketim_data = []
        islem_data = []

        for tarih_str in sorted(tum_gunler.keys()):
            # Tarih formatla (DD/MM)
            tarih_obj = datetime.strptime(tarih_str, "%Y-%m-%d")
            labels.append(tarih_obj.strftime("%d/%m"))
            tuketim_data.append(tum_gunler[tarih_str]["tuketim"])
            islem_data.append(tum_gunler[tarih_str]["islem_sayisi"])

        return jsonify(
            {
                "success": True,
                "labels": labels,
                "datasets": [
                    {
                        "label": "Toplam Tüketim",
                        "data": tuketim_data,
                        "borderColor": "rgb(239, 68, 68)",
                        "backgroundColor": "rgba(239, 68, 68, 0.1)",
                        "tension": 0.3,
                    },
                    {
                        "label": "İşlem Sayısı",
                        "data": islem_data,
                        "borderColor": "rgb(59, 130, 246)",
                        "backgroundColor": "rgba(59, 130, 246, 0.1)",
                        "tension": 0.3,
                    },
                ],
            }
        )

    except Exception:
        return jsonify({"success": False, "error": "Sunucu hatasi olustu"}), 500


# ==========================
# AUDIT TRAIL ROUTE'LARI
# ==========================


@app.route("/sistem-yoneticisi/audit-trail")
@login_required
@role_required("sistem_yoneticisi", "admin")
def audit_trail():
    """Audit Trail - Denetim İzi Sayfası"""
    from models import AuditLog, Kullanici
    from datetime import datetime, timedelta

    # Sayfalama
    page = request.args.get("page", 1, type=int)
    per_page = 50

    # Filtreler
    kullanici_id = request.args.get("kullanici_id", type=int)
    islem_tipi = request.args.get("islem_tipi")
    tablo_adi = request.args.get("tablo_adi")
    tarih_baslangic = request.args.get("tarih_baslangic")
    tarih_bitis = request.args.get("tarih_bitis")

    # Base query (superadmin aktiviteleri hariç)
    query = AuditLog.query.filter(AuditLog.kullanici_rol != "superadmin")

    # Filtreleme
    if kullanici_id:
        query = query.filter_by(kullanici_id=kullanici_id)
    if islem_tipi:
        query = query.filter_by(islem_tipi=islem_tipi)
    if tablo_adi:
        query = query.filter_by(tablo_adi=tablo_adi)
    if tarih_baslangic:
        tarih_baslangic_dt = datetime.strptime(tarih_baslangic, "%Y-%m-%d")
        query = query.filter(AuditLog.islem_tarihi >= tarih_baslangic_dt)
    if tarih_bitis:
        tarih_bitis_dt = datetime.strptime(tarih_bitis, "%Y-%m-%d") + timedelta(days=1)
        query = query.filter(AuditLog.islem_tarihi < tarih_bitis_dt)

    # Sıralama ve sayfalama
    query = query.order_by(AuditLog.islem_tarihi.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # İstatistikler
    bugun = get_kktc_now().replace(hour=0, minute=0, second=0, microsecond=0)
    bu_hafta = bugun - timedelta(days=bugun.weekday())
    bu_ay = bugun.replace(day=1)

    stats = {
        "today": AuditLog.query.filter(AuditLog.islem_tarihi >= bugun).count(),
        "week": AuditLog.query.filter(AuditLog.islem_tarihi >= bu_hafta).count(),
        "month": AuditLog.query.filter(AuditLog.islem_tarihi >= bu_ay).count(),
    }

    # Filtre için kullanıcı listesi
    users = (
        Kullanici.query.filter_by(aktif=True).order_by(Kullanici.kullanici_adi).all()
    )

    # Filtre için tablo listesi (unique)
    tables = (
        db.session.query(AuditLog.tablo_adi)
        .distinct()
        .order_by(AuditLog.tablo_adi)
        .all()
    )
    tables = [t[0] for t in tables]

    return render_template(
        "sistem_yoneticisi/audit_trail.html",
        logs=pagination.items,
        pagination=pagination,
        users=users,
        tables=tables,
        stats=stats,
    )


@app.route("/sistem-yoneticisi/audit-trail/<int:log_id>")
@login_required
@role_required("sistem_yoneticisi", "admin")
def audit_trail_detail(log_id):
    """Audit Log Detay API"""
    from models import AuditLog

    log = AuditLog.query.get_or_404(log_id)

    return jsonify(
        {
            "id": log.id,
            "kullanici_id": log.kullanici_id,
            "kullanici_adi": log.kullanici_adi,
            "kullanici_rol": log.kullanici_rol,
            "islem_tipi": log.islem_tipi,
            "tablo_adi": log.tablo_adi,
            "kayit_id": log.kayit_id,
            "eski_deger": log.eski_deger,
            "yeni_deger": log.yeni_deger,
            "degisiklik_ozeti": log.degisiklik_ozeti,
            "http_method": log.http_method,
            "url": log.url,
            "endpoint": log.endpoint,
            "ip_adresi": log.ip_adresi,
            "user_agent": log.user_agent,
            "islem_tarihi": log.islem_tarihi.strftime("%d.%m.%Y %H:%M:%S"),
            "aciklama": log.aciklama,
            "basarili": log.basarili,
            "hata_mesaji": log.hata_mesaji,
        }
    )


@app.route("/sistem-yoneticisi/audit-trail/export")
@login_required
@role_required("sistem_yoneticisi", "admin")
def audit_trail_export():
    """Audit Trail Excel Export"""
    from models import AuditLog
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime

    # Filtreleri al
    kullanici_id = request.args.get("kullanici_id", type=int)
    islem_tipi = request.args.get("islem_tipi")
    tablo_adi = request.args.get("tablo_adi")
    tarih_baslangic = request.args.get("tarih_baslangic")
    tarih_bitis = request.args.get("tarih_bitis")

    # Query oluştur (superadmin aktiviteleri hariç)
    query = AuditLog.query.filter(AuditLog.kullanici_rol != "superadmin")

    if kullanici_id:
        query = query.filter_by(kullanici_id=kullanici_id)
    if islem_tipi:
        query = query.filter_by(islem_tipi=islem_tipi)
    if tablo_adi:
        query = query.filter_by(tablo_adi=tablo_adi)
    if tarih_baslangic:
        tarih_baslangic_dt = datetime.strptime(tarih_baslangic, "%Y-%m-%d")
        query = query.filter(AuditLog.islem_tarihi >= tarih_baslangic_dt)
    if tarih_bitis:
        tarih_bitis_dt = datetime.strptime(tarih_bitis, "%Y-%m-%d")
        query = query.filter(AuditLog.islem_tarihi <= tarih_bitis_dt)

    logs = query.order_by(AuditLog.islem_tarihi.desc()).limit(10000).all()

    # Excel oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Audit Trail"

    # Başlıklar
    headers = [
        "ID",
        "Tarih",
        "Kullanıcı",
        "Rol",
        "İşlem",
        "Tablo",
        "Kayıt ID",
        "Değişiklik",
        "IP",
        "URL",
        "Başarılı",
    ]

    # Başlık satırını formatla
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # Verileri ekle
    for row, log in enumerate(logs, 2):
        ws.cell(row=row, column=1, value=log.id)
        ws.cell(row=row, column=2, value=log.islem_tarihi.strftime("%d.%m.%Y %H:%M"))
        ws.cell(row=row, column=3, value=log.kullanici_adi)
        ws.cell(row=row, column=4, value=log.kullanici_rol)
        ws.cell(row=row, column=5, value=log.islem_tipi)
        ws.cell(row=row, column=6, value=log.tablo_adi)
        ws.cell(row=row, column=7, value=log.kayit_id)
        ws.cell(row=row, column=8, value=log.degisiklik_ozeti or "")
        ws.cell(row=row, column=9, value=log.ip_adresi or "")
        ws.cell(row=row, column=10, value=log.url or "")
        ws.cell(row=row, column=11, value="Evet" if log.basarili else "Hayır")

    # Sütun genişliklerini ayarla
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 50
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 40
    ws.column_dimensions["K"].width = 10

    # Excel dosyasını kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"audit_trail_{get_kktc_now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Audit export işlemini logla
    from utils.audit import audit_export

    audit_export("audit_logs", f"Excel export: {len(logs)} kayıt")

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def init_database():
    """Veritabanı ve tabloları otomatik kontrol et - GÜVENLİ MOD"""
    try:
        with app.app_context():
            # Sadece bağlantı testi yap, tablo oluşturma!
            # Production'da mevcut verilere dokunmamak için
            inspector = inspect(db.engine)
            existing_tables = inspector.get_table_names()

            if existing_tables:
                print(
                    f"✅ Veritabanı bağlantısı başarılı - {len(existing_tables)} tablo mevcut"
                )
                return True
            else:
                print("⚠️  Henüz tablo yok!")
                print("🔧 Lütfen 'python init_db.py' komutunu çalıştırın.")
                return False
    except Exception as e:
        print(f"❌ Veritabanı hatası: {e}")
        print()
        print("🔧 Lütfen 'python init_db.py' komutunu çalıştırın.")
        return False


# ============================================
# Kat Sorumlusu route'ları → routes/kat_sorumlusu_routes.py'ye taşındı
# ============================================

# ============================================
# QR KOD SİSTEMİ ROTALARI
# ============================================

# ============================================
# Tüm route'lar merkezi olarak register edildi (routes/__init__.py)
# ============================================

# ============================================
# SCHEDULER - Otomatik Görevler
# ============================================
from apscheduler.schedulers.background import BackgroundScheduler  # noqa: E402
from apscheduler.triggers.cron import CronTrigger  # noqa: E402
from utils.file_management_service import FileManagementService  # noqa: E402


def start_scheduler():
    """Zamanlanmış görevleri başlat"""
    # Debug modunda sadece child process'te (gerçek uygulama) çalıştır
    # WERKZEUG_RUN_MAIN='true' sadece child process'te set edilir
    is_reloader_process = os.environ.get("WERKZEUG_RUN_MAIN") is None
    if is_reloader_process and app.debug:
        return  # Ana reloader process'inde scheduler başlatma

    # Gunicorn multi-worker: sadece belirlenmiş worker'da scheduler başlat
    # gunicorn.conf.py post_fork hook'unda SCHEDULER_WORKER_ID=1 set edilmeli
    if os.environ.get("SERVER_SOFTWARE", "").startswith("gunicorn"):
        scheduler_worker = os.environ.get("SCHEDULER_WORKER_ID")
        if scheduler_worker is None:
            # Hook set edilmemişse scheduler atla - Celery Beat kullanılmalı
            logger.warning(
                "Gunicorn multi-worker: SCHEDULER_WORKER_ID set edilmedi, scheduler atlanıyor"
            )
            return

    scheduler = BackgroundScheduler()

    # Her gün saat 02:00'de eski dosyaları temizle
    scheduler.add_job(
        func=lambda: FileManagementService.cleanup_old_files(),
        trigger=CronTrigger(hour=2, minute=0),
        id="cleanup_old_files",
        name="Eski doluluk dosyalarını temizle",
        replace_existing=True,
    )

    # ML SYSTEM JOBS - Sadece ML_ENABLED=true ise çalışır
    ml_enabled = os.getenv("ML_ENABLED", "false").lower() == "true"

    if ml_enabled:
        # Sabah 08:00 - Akşam 20:00 arası her saat başı veri toplama
        scheduler.add_job(
            func=lambda: collect_ml_data(),
            trigger="cron",
            hour="8-20",  # 08:00, 09:00, ..., 20:00
            minute=0,
            id="ml_data_collection",
            name="ML Veri Toplama",
            replace_existing=True,
        )

        # Sabah 08:00 - Akşam 20:00 arası her saat başı anomali tespiti (30. dakikada)
        scheduler.add_job(
            func=lambda: detect_anomalies(),
            trigger="cron",
            hour="8-20",  # 08:00, 09:00, ..., 20:00
            minute=30,  # Veri toplamadan 30 dk sonra
            id="ml_anomaly_detection",
            name="ML Anomali Tespiti",
            replace_existing=True,
        )

        # Her gece yarısı model eğitimi
        ml_training_schedule = os.getenv(
            "ML_TRAINING_SCHEDULE", "0 0 * * *"
        )  # Cron format
        scheduler.add_job(
            func=lambda: train_ml_models(),
            trigger=CronTrigger.from_crontab(ml_training_schedule),
            id="ml_model_training",
            name="ML Model Eğitimi",
            replace_existing=True,
        )

        # Günde 2 kez stok bitiş kontrolü (sabah 9 ve akşam 6)
        scheduler.add_job(
            func=lambda: check_stock_depletion(),
            trigger="cron",
            hour="9,18",
            id="ml_stock_depletion_check",
            name="ML Stok Bitiş Kontrolü",
            replace_existing=True,
        )

        # Her gece 03:00'te eski alertleri temizle
        scheduler.add_job(
            func=lambda: cleanup_old_alerts(),
            trigger="cron",
            hour=3,
            minute=0,
            id="ml_alert_cleanup",
            name="ML Alert Temizleme",
            replace_existing=True,
        )

        # Her gece 04:00'te eski model versiyonlarını temizle
        scheduler.add_job(
            func=lambda: cleanup_old_models(),
            trigger="cron",
            hour=4,
            minute=0,
            id="ml_model_cleanup",
            name="ML Model Cleanup",
            replace_existing=True,
        )

        print("✅ ML Scheduler başlatıldı")
        print("   - Veri toplama: 08:00-20:00 arası her saat başı")
        print("   - Anomali tespiti: 08:30-20:30 arası her saat")
        print(f"   - Model eğitimi: {ml_training_schedule}")
        print("   - Stok bitiş kontrolü: Günde 2 kez (09:00, 18:00)")
        print("   - Alert temizleme: Her gece 03:00")
        print("   - Model cleanup: Her gece 04:00")

    scheduler.start()
    print("✅ Scheduler başlatıldı (Günlük dosya temizleme: 02:00)")


def collect_ml_data():
    """ML veri toplama job'u"""
    try:
        from utils.ml.data_collector import DataCollector

        with app.app_context():
            collector = DataCollector(db)
            collector.collect_all_metrics()
            # Eski metrikleri temizle (90 günden eski)
            collector.cleanup_old_metrics(days=90)
    except Exception as e:
        logger.error(f"❌ ML veri toplama hatası: {str(e)}")


def detect_anomalies():
    """ML anomali tespiti job'u"""
    try:
        from utils.ml.anomaly_detector import AnomalyDetector

        with app.app_context():
            detector = AnomalyDetector(db)
            detector.detect_all_anomalies()
    except Exception as e:
        logger.error(f"❌ ML anomali tespiti hatası: {str(e)}")


def train_ml_models():
    """ML model eğitimi job'u"""
    try:
        from utils.ml.model_trainer import ModelTrainer

        with app.app_context():
            trainer = ModelTrainer(db)
            trainer.train_all_models()
    except Exception as e:
        logger.error(f"❌ ML model eğitimi hatası: {str(e)}")


def check_stock_depletion():
    """Stok bitiş kontrolü job'u"""
    try:
        from utils.ml.metrics_calculator import MetricsCalculator

        with app.app_context():
            calculator = MetricsCalculator(db)
            calculator.check_stock_depletion_alerts()
    except Exception as e:
        logger.error(f"❌ Stok bitiş kontrolü hatası: {str(e)}")


def cleanup_old_alerts():
    """Eski alertleri temizle job'u"""
    try:
        from utils.ml.alert_manager import AlertManager

        with app.app_context():
            alert_manager = AlertManager(db)
            alert_manager.cleanup_old_alerts(days=90)
    except Exception as e:
        logger.error(f"❌ Alert temizleme hatası: {str(e)}")


def cleanup_old_models():
    """Eski model versiyonlarını temizle job'u"""
    try:
        from utils.ml.model_manager import ModelManager

        with app.app_context():
            model_manager = ModelManager(db)

            # Eski model versiyonlarını temizle (son 3 versiyon sakla)
            result = model_manager.cleanup_old_models(keep_versions=3)

            # Disk kullanımını kontrol et
            disk_info = model_manager._check_disk_space()

            # Disk kullanımı %90'ı geçtiyse alert oluştur
            if disk_info["percent"] > 90:
                logger.warning(
                    f"⚠️  DISK KULLANIMI YÜKSEK: {disk_info['percent']:.1f}% "
                    f"({disk_info['used_gb']:.2f}GB / {disk_info['total_gb']:.2f}GB)"
                )

                # ML Alert oluştur
                from models import MLAlert

                alert = MLAlert(
                    alert_type="stok_anomali",  # En yakın tip
                    severity="kritik",
                    entity_id=0,  # Sistem seviyesi
                    metric_value=disk_info["percent"],
                    expected_value=80.0,
                    deviation_percent=(disk_info["percent"] - 80.0) / 80.0 * 100,
                    message=f"ML model dizini disk kullanımı kritik seviyede: {disk_info['percent']:.1f}%",
                    suggested_action="Eski model dosyalarını manuel olarak temizleyin veya disk alanını artırın",
                    created_at=get_kktc_now(),
                )
                db.session.add(alert)
                db.session.commit()

            logger.info(
                f"✅ Model cleanup tamamlandı: "
                f"{result['deleted_count']} model silindi, "
                f"{result['freed_space_mb']:.2f}MB alan boşaltıldı"
            )

    except Exception as e:
        logger.error(f"❌ Model cleanup hatası: {str(e)}")


# Scheduler'ı başlat
try:
    start_scheduler()
except Exception as e:
    print(f"⚠️  Scheduler başlatılamadı: {str(e)}")

# ============================================


# Error Handlers - Session timeout ve diğer hatalar için
@app.errorhandler(500)
def internal_error(error):
    """500 hatası - Session timeout durumunda login'e yönlendir"""
    db.session.rollback()
    if "kullanici_id" not in session:
        flash("Oturumunuz sona erdi. Lütfen tekrar giriş yapın.", "warning")
        return redirect(url_for("login"))
    logger.error(f"500 Hatası: {error}")
    return render_template("errors/500.html"), 500


@app.errorhandler(401)
def unauthorized_error(error):
    """401 hatası - Yetkisiz erişim"""
    flash("Bu sayfaya erişim yetkiniz yok. Lütfen giriş yapın.", "warning")
    return redirect(url_for("login"))


@app.errorhandler(403)
def forbidden_error(error):
    """403 hatası - Yasaklı erişim"""
    flash("Bu sayfaya erişim yetkiniz yok.", "danger")
    return redirect(url_for("dashboard"))


@app.errorhandler(404)
def not_found_error(error):
    """404 hatası - Sayfa bulunamadı"""
    if "kullanici_id" not in session:
        return redirect(url_for("login"))
    return render_template("errors/404.html"), 404


# Session kontrolü - Her istekte
@app.before_request
def check_session_validity():
    """Her istekte session geçerliliğini kontrol et"""
    from datetime import datetime as dt

    # Static dosyalar ve login sayfası hariç
    if (
        request.endpoint
        and not request.endpoint.startswith("static")
        and request.endpoint not in ["login", "logout"]
    ):
        if "kullanici_id" in session:
            # Session son aktivite zamanını kontrol et
            last_activity = session.get("last_activity")
            if last_activity:
                try:
                    last_time = dt.fromisoformat(last_activity)
                    # Timezone-aware karşılaştırma (KKTC)
                    if last_time.tzinfo is None:
                        last_time = KKTC_TZ.localize(last_time)
                    timeout = app.config.get(
                        "PERMANENT_SESSION_LIFETIME", timedelta(hours=8)
                    )
                    if isinstance(timeout, int):
                        timeout = timedelta(seconds=timeout)
                    if get_kktc_now() - last_time > timeout:
                        session.clear()
                        flash(
                            "Oturumunuz sona erdi. Lütfen tekrar giriş yapın.",
                            "warning",
                        )
                        return redirect(url_for("login"))
                except Exception:
                    logger.debug("Sessiz hata yakalandi", exc_info=True)
            # Son aktivite zamanını güncelle (timezone-aware)
            session["last_activity"] = get_kktc_now().isoformat()


if __name__ == "__main__":
    print()
    print("=" * 60)
    print("🏨 OTEL MİNİBAR TAKİP SİSTEMİ")
    print("=" * 60)
    print()

    # Veritabanını başlat
    if init_database():
        print()
        print("🚀 Uygulama başlatılıyor...")
        # Railway PORT environment variable desteği
        port = int(os.getenv("PORT", 5014))
        debug_mode = os.getenv("FLASK_ENV", "development") == "development"

        # HTTPS desteği için SSL context (mobil kamera erişimi için gerekli)
        ssl_context = None
        use_https = os.getenv("USE_HTTPS", "false").lower() == "true"

        if use_https:
            cert_file = os.path.join(os.path.dirname(__file__), "cert.pem")
            key_file = os.path.join(os.path.dirname(__file__), "key.pem")

            if os.path.exists(cert_file) and os.path.exists(key_file):
                ssl_context = (cert_file, key_file)
                print(f"🔒 HTTPS Aktif: https://0.0.0.0:{port}")
                print(f"📱 Mobil erişim: https://<IP-ADRESINIZ>:{port}")
                print(
                    "⚠️  Self-signed sertifika kullanıldığı için tarayıcıda güvenlik uyarısı alabilirsiniz."
                )
                print("   Mobilde 'Advanced' > 'Proceed to site' seçeneğini kullanın.")
            else:
                print("⚠️  SSL sertifikası bulunamadı. Sertifika oluşturmak için:")
                print("   python generate_ssl_cert.py")
                print("📍 HTTP Modu: http://0.0.0.0:{port}")
                print("⚠️  Mobil kamera erişimi için HTTPS gereklidir!")
        else:
            print(f"📍 HTTP Modu: http://0.0.0.0:{port}")
            print("⚠️  Mobil kamera erişimi için HTTPS gereklidir!")
            print(
                "   HTTPS'i aktifleştirmek için .env dosyasına USE_HTTPS=true ekleyin"
            )

        print("🌙 Dark/Light tema: Sağ üstten değiştirilebilir")
        print()
        print("Durdurmak için CTRL+C kullanın")
        print("=" * 60)
        print()

        try:
            app.run(
                debug=debug_mode,
                host="0.0.0.0",
                port=port,
                ssl_context=ssl_context,
                use_reloader=True,
            )
        except Exception as e:
            print(f"❌ Flask başlatma hatası: {e}")
            import traceback

            traceback.print_exc()
    else:
        print()
        print("❌ Uygulama başlatılamadı. Lütfen veritabanı ayarlarını kontrol edin.")
        print()
        exit(1)


# ============================================================================
# API: SETUP YÖNETİMİ → routes/sistem_yoneticisi_routes.py'de tanımlı
# ============================================================================


# ============================================================================
# DOCKER HEALTH CHECK ENDPOINT
# ============================================================================

@app.route('/health', methods=['GET'])
def health_check():
    """
    Docker container health check endpoint
    Database bağlantısını kontrol eder
    """
    try:
        # Database bağlantısını test et
        db.session.execute(db.text('SELECT 1'))
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'timestamp': get_kktc_now().isoformat()
        }), 200
    except Exception:
        return jsonify(
            {
                "status": "unhealthy",
                "database": "disconnected",
                "error": "Sunucu hatasi olustu",
                "timestamp": get_kktc_now().isoformat(),
            }
        ), 503
