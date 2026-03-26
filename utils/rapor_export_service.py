"""
Rapor Export Servisi

Excel ve PDF rapor uretme mantigi.
app.py'deki ~600 satirlik god function'lardan cikarilmistir.
Celery task'lar ve senkron route'lar bu servis uzerinden calisir.
"""

import io
import logging
from datetime import datetime, timedelta

import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from models import (
    db,
    Urun,
    UrunGrup,
    StokHareket,
    PersonelZimmet,
    PersonelZimmetDetay,
    Kullanici,
)
from utils.helpers import get_stok_toplamlari, get_kritik_stok_urunler

logger = logging.getLogger(__name__)

KKTC_TZ = pytz.timezone("Europe/Nicosia")

RAPOR_BASLIKLARI = {
    "stok_durum": "Stok Durum Raporu",
    "stok_hareket": "Stok Hareket Raporu",
    "zimmet": "Zimmet Raporu",
    "zimmet_detay": "Ürün Bazlı Zimmet Detay Raporu",
    "urun_grup": "Ürün Grubu Bazlı Stok Raporu",
    "ozet": "Genel Sistem Özet Raporu",
}

VALID_RAPOR_TIPLERI = set(RAPOR_BASLIKLARI.keys())


def _get_kktc_now():
    return datetime.now(KKTC_TZ)


def _parse_date(date_str) -> datetime | None:
    if not date_str:
        return None
    return datetime.strptime(date_str, "%Y-%m-%d")


def _turkce_ascii(text):
    """Turkce karakterleri ASCII'ye donustur (PDF font uyumlulugu icin)."""
    if not text:
        return ""
    char_map = {
        "ç": "c",
        "Ç": "C",
        "ğ": "g",
        "Ğ": "G",
        "ı": "i",
        "İ": "I",
        "ö": "o",
        "Ö": "O",
        "ş": "s",
        "Ş": "S",
        "ü": "u",
        "Ü": "U",
    }
    result = str(text)
    for turkish, ascii_char in char_map.items():
        result = result.replace(turkish, ascii_char)
    return result


def _auto_column_width(ws):
    """Excel sutun genisliklerini icerikteki en uzun degere gore ayarla."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _write_excel_headers(ws, row_num, headers, color="4472C4"):
    """Baslik satirini yaz ve stil uygula."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def _apply_date_filters_stok_hareket(query, filters):
    """StokHareket sorgusuna tarih ve urun filtrelerini uygula."""
    if filters.get("baslangic_tarihi"):
        baslangic = _parse_date(filters["baslangic_tarihi"])
        query = query.filter(StokHareket.islem_tarihi >= baslangic)
    if filters.get("bitis_tarihi"):
        bitis = _parse_date(filters["bitis_tarihi"]) + timedelta(days=1)  # type: ignore[operator]
        query = query.filter(StokHareket.islem_tarihi < bitis)
    if filters.get("urun_id"):
        query = query.filter_by(urun_id=filters["urun_id"])
    elif filters.get("urun_grup_id"):
        query = query.join(Urun).filter(Urun.grup_id == filters["urun_grup_id"])
    if filters.get("hareket_tipi"):
        query = query.filter_by(hareket_tipi=filters["hareket_tipi"])
    return query


def _apply_date_filters_zimmet(query, filters):
    """PersonelZimmet sorgusuna tarih ve personel filtrelerini uygula."""
    if filters.get("baslangic_tarihi"):
        baslangic = _parse_date(filters["baslangic_tarihi"])
        query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
    if filters.get("bitis_tarihi"):
        bitis = _parse_date(filters["bitis_tarihi"]) + timedelta(days=1)  # type: ignore[operator]
        query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
    if filters.get("personel_id"):
        query = query.filter_by(personel_id=filters["personel_id"])
    return query


def _enrich_zimmet_aciklama(hareket):
    """Stok hareket aciklamasina zimmet personel bilgisini ekle."""
    aciklama = hareket.aciklama or "-"
    if hareket.aciklama and "Zimmet" in hareket.aciklama:
        try:
            if "#" in hareket.aciklama:
                zimmet_id = int(hareket.aciklama.split("#")[1].split()[0])
                zimmet = db.session.get(PersonelZimmet, zimmet_id)
                if zimmet and zimmet.personel:
                    aciklama += f" → {zimmet.personel.ad} {zimmet.personel.soyad}"
        except Exception:
            logger.debug("Zimmet bilgisi eklenemedi", exc_info=True)
    return aciklama


def _get_urun_grup_stats():
    """Urun grubu bazli stok istatistiklerini hesapla."""
    gruplar = UrunGrup.query.filter_by(aktif=True).all()
    aktif_urunler = Urun.query.filter_by(aktif=True).all()
    stok_map = (
        get_stok_toplamlari([u.id for u in aktif_urunler]) if aktif_urunler else {}
    )

    grup_urunleri_map = {}
    for urun in aktif_urunler:
        grup_urunleri_map.setdefault(urun.grup_id, []).append(urun)

    results = []
    for grup in gruplar:
        grup_urunleri = grup_urunleri_map.get(grup.id, [])
        toplam = len(grup_urunleri)
        kritik = sum(
            1
            for u in grup_urunleri
            if stok_map.get(u.id, 0) <= (u.kritik_stok_seviyesi or 0)
        )
        results.append((grup, toplam, kritik))
    return results


# ============================================================================
# EXCEL REPORT GENERATORS
# ============================================================================


def _excel_stok_durum(ws, row_num, filters):
    headers = [
        "Ürün Adı",
        "Ürün Grubu",
        "Birim",
        "Mevcut Stok",
        "Kritik Seviye",
        "Durum",
    ]
    ws.merge_cells("A1:F1")
    _write_excel_headers(ws, row_num, headers)

    query = Urun.query.filter_by(aktif=True)
    if filters.get("urun_grup_id"):
        query = query.filter_by(grup_id=filters["urun_grup_id"])

    urunler_liste = query.order_by(Urun.urun_adi).all()
    stok_map = get_stok_toplamlari([u.id for u in urunler_liste])

    for urun in urunler_liste:
        row_num += 1
        mevcut_stok = stok_map.get(urun.id, 0)
        kritik_seviye = urun.kritik_stok_seviyesi or 0
        durum = "KRİTİK" if mevcut_stok <= kritik_seviye else "NORMAL"

        ws.cell(row=row_num, column=1, value=urun.urun_adi)
        ws.cell(row=row_num, column=2, value=urun.grup.grup_adi)
        ws.cell(row=row_num, column=3, value=urun.birim)
        ws.cell(row=row_num, column=4, value=mevcut_stok)
        ws.cell(row=row_num, column=5, value=urun.kritik_stok_seviyesi)
        ws.cell(row=row_num, column=6, value=durum)
    return row_num


def _excel_stok_hareket(ws, row_num, filters):
    headers = ["Tarih", "Ürün Adı", "Hareket Tipi", "Miktar", "Açıklama", "İşlem Yapan"]
    ws.merge_cells("A1:F1")
    _write_excel_headers(ws, row_num, headers, "70AD47")

    query = _apply_date_filters_stok_hareket(StokHareket.query, filters)
    hareketler = query.order_by(StokHareket.islem_tarihi.desc()).limit(10000).all()

    for hareket in hareketler:
        row_num += 1
        islem_yapan = (
            f"{hareket.islem_yapan.ad} {hareket.islem_yapan.soyad}"
            if hareket.islem_yapan
            else "-"
        )
        aciklama = _enrich_zimmet_aciklama(hareket)

        ws.cell(
            row=row_num, column=1, value=hareket.islem_tarihi.strftime("%d.%m.%Y %H:%M")
        )
        ws.cell(row=row_num, column=2, value=hareket.urun.urun_adi)
        ws.cell(row=row_num, column=3, value=hareket.hareket_tipi.upper())
        ws.cell(row=row_num, column=4, value=hareket.miktar)
        ws.cell(row=row_num, column=5, value=aciklama)
        ws.cell(row=row_num, column=6, value=islem_yapan)
    return row_num


def _excel_zimmet(ws, row_num, filters):
    headers = [
        "Zimmet No",
        "Personel",
        "Zimmet Tarihi",
        "Ürün Sayısı",
        "Toplam Miktar",
        "Durum",
    ]
    ws.merge_cells("A1:F1")
    _write_excel_headers(ws, row_num, headers, "FFC000")

    query = _apply_date_filters_zimmet(PersonelZimmet.query, filters)
    zimmetler = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).limit(10000).all()

    for zimmet in zimmetler:
        row_num += 1
        toplam_miktar = sum(d.miktar for d in zimmet.detaylar)

        ws.cell(row=row_num, column=1, value=f"#{zimmet.id}")
        ws.cell(
            row=row_num, column=2, value=f"{zimmet.personel.ad} {zimmet.personel.soyad}"
        )
        ws.cell(
            row=row_num, column=3, value=zimmet.zimmet_tarihi.strftime("%d.%m.%Y %H:%M")
        )
        ws.cell(row=row_num, column=4, value=len(zimmet.detaylar))
        ws.cell(row=row_num, column=5, value=toplam_miktar)
        ws.cell(row=row_num, column=6, value=zimmet.durum.upper())
    return row_num


def _excel_zimmet_detay(ws, row_num, filters):
    headers = [
        "Zimmet No",
        "Personel",
        "Zimmet Tarihi",
        "Ürün Adı",
        "Grup",
        "Miktar",
        "Durum",
    ]
    ws.merge_cells("A1:G1")
    _write_excel_headers(ws, row_num, headers, "C55A11")

    query = (
        db.session.query(PersonelZimmetDetay, PersonelZimmet, Kullanici, Urun)
        .join(PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id)
        .join(Kullanici, PersonelZimmet.personel_id == Kullanici.id)
        .join(Urun, PersonelZimmetDetay.urun_id == Urun.id)
    )

    if filters.get("baslangic_tarihi"):
        baslangic = _parse_date(filters["baslangic_tarihi"])
        query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
    if filters.get("bitis_tarihi"):
        bitis = _parse_date(filters["bitis_tarihi"]) + timedelta(days=1)  # type: ignore[operator]
        query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
    if filters.get("personel_id"):
        query = query.filter(PersonelZimmet.personel_id == filters["personel_id"])
    if filters.get("urun_id"):
        query = query.filter(PersonelZimmetDetay.urun_id == filters["urun_id"])
    elif filters.get("urun_grup_id"):
        query = query.filter(Urun.grup_id == filters["urun_grup_id"])

    detaylar = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).limit(10000).all()

    for detay, zimmet, kullanici, urun in detaylar:
        row_num += 1
        ws.cell(row=row_num, column=1, value=f"#{zimmet.id}")
        ws.cell(row=row_num, column=2, value=f"{kullanici.ad} {kullanici.soyad}")
        ws.cell(
            row=row_num, column=3, value=zimmet.zimmet_tarihi.strftime("%d.%m.%Y %H:%M")
        )
        ws.cell(row=row_num, column=4, value=urun.urun_adi)
        ws.cell(row=row_num, column=5, value=urun.grup.grup_adi)
        ws.cell(row=row_num, column=6, value=detay.miktar)
        ws.cell(row=row_num, column=7, value=zimmet.durum.upper())
    return row_num


def _excel_urun_grup(ws, row_num, filters):
    headers = ["Ürün Grubu", "Toplam Ürün", "Kritik Stoklu Ürün"]
    ws.merge_cells("A1:C1")
    _write_excel_headers(ws, row_num, headers, "5B9BD5")

    for grup, toplam, kritik in _get_urun_grup_stats():
        row_num += 1
        ws.cell(row=row_num, column=1, value=grup.grup_adi)
        ws.cell(row=row_num, column=2, value=toplam)
        ws.cell(row=row_num, column=3, value=kritik)
    return row_num


_EXCEL_GENERATORS = {
    "stok_durum": _excel_stok_durum,
    "stok_hareket": _excel_stok_hareket,
    "zimmet": _excel_zimmet,
    "zimmet_detay": _excel_zimmet_detay,
    "urun_grup": _excel_urun_grup,
}


def generate_excel(rapor_tipi, filters=None):
    """Excel raporu uret ve BytesIO dondur.

    Args:
        rapor_tipi: 'stok_durum', 'stok_hareket', 'zimmet', 'zimmet_detay', 'urun_grup'
        filters: dict - baslangic_tarihi, bitis_tarihi, urun_grup_id, urun_id, personel_id, hareket_tipi

    Returns:
        io.BytesIO
    """
    if rapor_tipi not in VALID_RAPOR_TIPLERI:
        raise ValueError(f"Gecersiz rapor tipi: {rapor_tipi}")

    filters = filters or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    baslik = RAPOR_BASLIKLARI.get(rapor_tipi, "Rapor")
    ws.title = baslik[:31]

    ws["A1"] = baslik
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Rapor Tarihi: {_get_kktc_now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(size=10)

    row_num = 4

    generator = _EXCEL_GENERATORS.get(rapor_tipi)
    if generator:
        generator(ws, row_num, filters)

    _auto_column_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================================
# PDF REPORT GENERATORS
# ============================================================================


def _pdf_stok_durum(filters):
    data = [
        [
            _turkce_ascii(h)
            for h in [
                "Urun Adi",
                "Urun Grubu",
                "Birim",
                "Mevcut Stok",
                "Kritik Seviye",
                "Durum",
            ]
        ]
    ]

    query = Urun.query.filter_by(aktif=True)
    if filters.get("urun_grup_id"):
        query = query.filter_by(grup_id=filters["urun_grup_id"])

    urunler_liste = query.order_by(Urun.urun_adi).all()
    stok_map = (
        get_stok_toplamlari([u.id for u in urunler_liste]) if urunler_liste else {}
    )

    for urun in urunler_liste:
        mevcut_stok = stok_map.get(urun.id, 0)
        kritik_seviye = urun.kritik_stok_seviyesi or 0
        durum = "KRITIK" if mevcut_stok <= kritik_seviye else "NORMAL"
        data.append(
            [
                _turkce_ascii(urun.urun_adi),
                _turkce_ascii(urun.grup.grup_adi),
                _turkce_ascii(urun.birim),
                str(mevcut_stok),
                str(urun.kritik_stok_seviyesi),
                durum,
            ]
        )
    return data


def _pdf_stok_hareket(filters):
    data = [
        [
            _turkce_ascii(h)
            for h in ["Tarih", "Urun Adi", "Hareket", "Miktar", "Aciklama"]
        ]
    ]

    query = _apply_date_filters_stok_hareket(StokHareket.query, filters)
    hareketler = query.order_by(StokHareket.islem_tarihi.desc()).limit(100).all()

    for hareket in hareketler:
        aciklama = _enrich_zimmet_aciklama(hareket)
        aciklama_kisaltilmis = aciklama[:50] if len(aciklama) > 50 else aciklama
        data.append(
            [
                hareket.islem_tarihi.strftime("%d.%m.%Y %H:%M"),
                _turkce_ascii(hareket.urun.urun_adi),
                _turkce_ascii(hareket.hareket_tipi.upper()),
                str(hareket.miktar),
                _turkce_ascii(aciklama_kisaltilmis),
            ]
        )
    return data


def _pdf_zimmet(filters):
    data = [
        [
            _turkce_ascii(h)
            for h in [
                "Zimmet No",
                "Personel",
                "Tarih",
                "Urun Sayisi",
                "Toplam",
                "Durum",
            ]
        ]
    ]

    query = _apply_date_filters_zimmet(PersonelZimmet.query, filters)
    zimmetler = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).limit(100).all()

    for zimmet in zimmetler:
        toplam_miktar = sum(d.miktar for d in zimmet.detaylar)
        data.append(
            [
                f"#{zimmet.id}",
                _turkce_ascii(f"{zimmet.personel.ad} {zimmet.personel.soyad}"),
                zimmet.zimmet_tarihi.strftime("%d.%m.%Y"),
                str(len(zimmet.detaylar)),
                str(toplam_miktar),
                _turkce_ascii(zimmet.durum.upper()),
            ]
        )
    return data


def _pdf_zimmet_detay(filters):
    data = [
        [
            _turkce_ascii(h)
            for h in ["Zimmet", "Personel", "Urun", "Grup", "Miktar", "Durum"]
        ]
    ]

    query = (
        db.session.query(PersonelZimmetDetay, PersonelZimmet, Kullanici, Urun)
        .join(PersonelZimmet, PersonelZimmetDetay.zimmet_id == PersonelZimmet.id)
        .join(Kullanici, PersonelZimmet.personel_id == Kullanici.id)
        .join(Urun, PersonelZimmetDetay.urun_id == Urun.id)
    )

    if filters.get("baslangic_tarihi"):
        baslangic = _parse_date(filters["baslangic_tarihi"])
        query = query.filter(PersonelZimmet.zimmet_tarihi >= baslangic)
    if filters.get("bitis_tarihi"):
        bitis = _parse_date(filters["bitis_tarihi"]) + timedelta(days=1)  # type: ignore[operator]
        query = query.filter(PersonelZimmet.zimmet_tarihi < bitis)
    if filters.get("personel_id"):
        query = query.filter(PersonelZimmet.personel_id == filters["personel_id"])
    if filters.get("urun_id"):
        query = query.filter(PersonelZimmetDetay.urun_id == filters["urun_id"])
    elif filters.get("urun_grup_id"):
        query = query.filter(Urun.grup_id == filters["urun_grup_id"])

    detaylar = query.order_by(PersonelZimmet.zimmet_tarihi.desc()).limit(100).all()

    for detay, zimmet, kullanici, urun in detaylar:
        data.append(
            [
                f"#{zimmet.id}",
                _turkce_ascii(f"{kullanici.ad} {kullanici.soyad}"),
                _turkce_ascii(urun.urun_adi),
                _turkce_ascii(urun.grup.grup_adi),
                str(detay.miktar),
                _turkce_ascii(zimmet.durum.upper()),
            ]
        )
    return data


def _pdf_urun_grup(filters):
    data = [
        [_turkce_ascii(h) for h in ["Urun Grubu", "Toplam Urun", "Kritik Stoklu Urun"]]
    ]

    for grup, toplam, kritik in _get_urun_grup_stats():
        data.append([_turkce_ascii(grup.grup_adi), str(toplam), str(kritik)])
    return data


def _pdf_ozet(filters):
    toplam_urun = Urun.query.filter_by(aktif=True).count()
    kritik_urunler = get_kritik_stok_urunler()
    aktif_zimmet = PersonelZimmet.query.filter_by(durum="aktif").count()

    bugun = _get_kktc_now().date()
    bugun_baslangic = datetime.combine(bugun, datetime.min.time())
    bugun_bitis = datetime.combine(bugun, datetime.max.time())

    bugun_giris = StokHareket.query.filter(
        StokHareket.hareket_tipi == "giris",
        StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis),
    ).count()

    bugun_cikis = StokHareket.query.filter(
        StokHareket.hareket_tipi == "cikis",
        StokHareket.islem_tarihi.between(bugun_baslangic, bugun_bitis),
    ).count()

    ay_baslangic = _get_kktc_now().replace(
        day=1, hour=0, minute=0, second=0, microsecond=0
    )
    ay_zimmet = PersonelZimmet.query.filter(
        PersonelZimmet.zimmet_tarihi >= ay_baslangic
    ).count()

    return [
        ["Metrik", "Deger"],
        [_turkce_ascii("Toplam Urun Sayisi"), str(toplam_urun)],
        [_turkce_ascii("Kritik Stoklu Urun"), str(len(kritik_urunler))],
        [_turkce_ascii("Aktif Zimmet"), str(aktif_zimmet)],
        [_turkce_ascii("Bugun Stok Giris"), str(bugun_giris)],
        [_turkce_ascii("Bugun Stok Cikis"), str(bugun_cikis)],
        [_turkce_ascii("Bu Ay Zimmet"), str(ay_zimmet)],
    ]


_PDF_GENERATORS = {
    "stok_durum": _pdf_stok_durum,
    "stok_hareket": _pdf_stok_hareket,
    "zimmet": _pdf_zimmet,
    "zimmet_detay": _pdf_zimmet_detay,
    "urun_grup": _pdf_urun_grup,
    "ozet": _pdf_ozet,
}


def generate_pdf(rapor_tipi, filters=None):
    """PDF raporu uret ve BytesIO dondur.

    Args:
        rapor_tipi: 'stok_durum', 'stok_hareket', 'zimmet', 'zimmet_detay', 'urun_grup', 'ozet'
        filters: dict - baslangic_tarihi, bitis_tarihi, urun_grup_id, urun_id, personel_id, hareket_tipi

    Returns:
        io.BytesIO
    """
    if rapor_tipi not in VALID_RAPOR_TIPLERI:
        raise ValueError(f"Gecersiz rapor tipi: {rapor_tipi}")

    filters = filters or {}

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    story = []

    baslik = _turkce_ascii(RAPOR_BASLIKLARI.get(rapor_tipi, "Rapor"))
    story.append(Paragraph(baslik, styles["Title"]))
    story.append(Spacer(1, 12))

    date_text = f"Rapor Tarihi: {_get_kktc_now().strftime('%d.%m.%Y %H:%M')}"
    story.append(Paragraph(date_text, styles["Normal"]))
    story.append(Spacer(1, 20))

    generator = _PDF_GENERATORS.get(rapor_tipi)
    data = generator(filters) if generator else []

    if data:
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("TOPPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#F2F2F2")],
                    ),
                ]
            )
        )
        story.append(table)
    else:
        story.append(
            Paragraph(
                _turkce_ascii("Bu filtre kriterleri icin veri bulunamadi."),
                styles["Normal"],
            )
        )

    doc.build(story)
    buffer.seek(0)
    return buffer


def get_export_filename(rapor_tipi, extension):
    """Dosya adini olustur: rapor_tipi_raporu_YYYYMMDD_HHMMSS.ext"""
    timestamp = _get_kktc_now().strftime("%Y%m%d_%H%M%S")
    return f"{rapor_tipi}_raporu_{timestamp}.{extension}"
